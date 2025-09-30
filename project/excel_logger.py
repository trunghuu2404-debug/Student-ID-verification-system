import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import cv2
from PIL import Image
import io


# For the project, we only want to save date,labels and ust image annotation into the excel file
# for the failed_verification excel file, there will be 1 additional column that mentioned
# the reason it failed
class VerificationLogger:
    def __init__(self):
        self.success_file = "successful_verifications.xlsx"
        self.failure_file = "failed_verifications.xlsx"
        self.max_image_height = 180  # Maximum height for Excel images
        self.max_image_width = 240  # Maximum width for Excel images

        # Create success file if it doesn't exist
        if not os.path.exists(self.success_file):
            self._create_success_file()

        # Create failure file if it doesn't exist
        if not os.path.exists(self.failure_file):
            self._create_failure_file()

    def _create_success_file(self):
        """Create the successful verifications Excel file with headers"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Verifications"

        # Define headers
        headers = [
            "Timestamp",
            "Student ID",
            "First Name",
            "Last Name",
            "Face Match",
            "Logo Found",
            "Pattern Count",
            "OTP Verified",
            "Annotated Image",
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="00B050", end_color="00B050", fill_type="solid"
            )
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Make the image column wider
        ws.column_dimensions[get_column_letter(len(headers))].width = 35
        # Make row height taller for images
        ws.row_dimensions[1].height = self.max_image_height + 20

        wb.save(self.success_file)

    def _create_failure_file(self):
        """Create the failed verifications Excel file with headers"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed Verifications"

        # Define headers
        headers = [
            "Timestamp",
            "Student ID",
            "First Name",
            "Last Name",
            "Face Match",
            "Logo Found",
            "Pattern Count",
            "Failure Reasons",
            "Annotated Image",
        ]

        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            )
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Make the Failure Reasons and image columns wider
        ws.column_dimensions[get_column_letter(len(headers) - 1)].width = 40
        ws.column_dimensions[get_column_letter(len(headers))].width = 35
        # Make row height taller for images
        ws.row_dimensions[1].height = self.max_image_height + 20

        wb.save(self.failure_file)

    def _resize_image_for_excel(self, image_array):
        """Resize image to fit Excel cell while maintaining aspect ratio"""
        # Convert CV2 image to PIL Image
        image = cv2.cvtColor(image_array, cv2.COLOR_BGR2RGB)
        pil_image = Image.fromarray(image)

        # Calculate new dimensions maintaining aspect ratio
        width, height = pil_image.size
        aspect_ratio = width / height

        if height > self.max_image_height:
            new_height = self.max_image_height
            new_width = int(new_height * aspect_ratio)
        else:
            new_width = width
            new_height = height

        if new_width > self.max_image_width:
            new_width = self.max_image_width
            new_height = int(new_width / aspect_ratio)

        # Resize image
        pil_image = pil_image.resize((new_width, new_height), Image.Resampling.LANCZOS)

        return pil_image, new_width, new_height

    def _add_image_to_cell(self, ws, row, col, image_array):
        """Add resized image to Excel cell"""
        # Resize image
        pil_image, width, height = self._resize_image_for_excel(image_array)

        # Save to bytes buffer
        img_bytes = io.BytesIO()
        pil_image.save(img_bytes, format="PNG")
        img_bytes.seek(0)

        # Add to Excel
        img = XLImage(img_bytes)
        img.width = width
        img.height = height

        # Position image in cell
        ws.row_dimensions[row].height = height + 20
        img.anchor = ws.cell(row=row, column=col).coordinate

        return img

    def log_successful_verification(
        self, verification_data, annotated_image, otp_verified=True
    ):
        """Log a successful verification with image"""
        try:
            wb = load_workbook(self.success_file)
            ws = wb.active

            # Get next empty row
            next_row = ws.max_row + 1

            # Add verification data
            ws.cell(
                row=next_row,
                column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
            ws.cell(
                row=next_row, column=2, value=verification_data.get("id_number", "N/A")
            )
            ws.cell(
                row=next_row, column=3, value=verification_data.get("first_name", "N/A")
            )
            ws.cell(
                row=next_row, column=4, value=verification_data.get("last_name", "N/A")
            )
            ws.cell(
                row=next_row,
                column=5,
                value=verification_data.get("face_match_result", "N/A"),
            )
            ws.cell(
                row=next_row,
                column=6,
                value=str(verification_data.get("logo_found", False)),
            )
            ws.cell(
                row=next_row, column=7, value=verification_data.get("pattern_count", 0)
            )
            ws.cell(row=next_row, column=8, value=str(otp_verified))

            # Add image to the last column
            img = self._add_image_to_cell(ws, next_row, 9, annotated_image)
            ws.add_image(img)

            wb.save(self.success_file)
            return True
        except Exception as e:
            print(f"Error logging successful verification: {e}")
            return False

    def log_failed_verification(self, verification_data, annotated_image):
        """Log a failed verification with image"""
        try:
            wb = load_workbook(self.failure_file)
            ws = wb.active

            # Get next empty row
            next_row = ws.max_row + 1

            # Add verification data
            ws.cell(
                row=next_row,
                column=1,
                value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
            ws.cell(
                row=next_row, column=2, value=verification_data.get("id_number", "N/A")
            )
            ws.cell(
                row=next_row, column=3, value=verification_data.get("first_name", "N/A")
            )
            ws.cell(
                row=next_row, column=4, value=verification_data.get("last_name", "N/A")
            )
            ws.cell(
                row=next_row,
                column=5,
                value=verification_data.get("face_match_result", "N/A"),
            )
            ws.cell(
                row=next_row,
                column=6,
                value=str(verification_data.get("logo_found", False)),
            )
            ws.cell(
                row=next_row, column=7, value=verification_data.get("pattern_count", 0)
            )

            # Combine all failure reasons into one string
            failure_reasons = " | ".join(
                verification_data.get("failure_reasons", ["Unknown"])
            )
            ws.cell(row=next_row, column=8, value=failure_reasons)

            # Add image to the last column
            img = self._add_image_to_cell(ws, next_row, 9, annotated_image)
            ws.add_image(img)

            wb.save(self.failure_file)
            return True
        except Exception as e:
            print(f"Error logging failed verification: {e}")
            return False
